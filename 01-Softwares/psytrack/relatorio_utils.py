from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import calendar
import os
import sys
import flet as ft


def get_path(relative_path):
    """Retorna o caminho absoluto mesmo dentro do .exe"""
    if getattr(sys, 'frozen', False):
        # Estamos rodando dentro do .exe (empacotado)
        base_path = sys._MEIPASS
    else:
        # Modo de desenvolvimento
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def gerar_relatorio_excel(dados, psicologo, mes, ano):
    if not dados:
        ft.AlertDialog(title=ft.Text("Nenhum dado encontrado")).open = True
        return None

    nome_mes = calendar.month_name[mes].capitalize()
    path_modelo = str(get_path("modelo/planilha_atendimento.xlsx"))
    path_saida = os.path.join(f"relatorios/relatorio_{psicologo}_{nome_mes}_{ano}.xlsx")

    wb = load_workbook(path_modelo)
    ws = wb.active

    # Mescla título
    ws["A2"].value = f"Psicólogo(a): {psicologo}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Inserção de dados a partir da linha 5, coluna 2
    row_start = 5
    for idx, sessao in enumerate(dados):
        ws.cell(row=row_start + idx, column=2, value=sessao[1])  # Nome do paciente
        ws.cell(row=row_start + idx, column=3, value=nome_mes)   # Mês de referência
        ws.cell(row=row_start + idx, column=4, value=sessao[2])  # Data
        ws.cell(row=row_start + idx, column=5, value=sessao[3])  # Realizada
        # Cobrar: 1 ou 0 com base na justificativa
        cobrar = 1 if "Cobrar" in sessao[4] and "Não Cobrar" not in sessao[4] else 0
        ws.cell(row=row_start + idx, column=6, value=cobrar)
        ws.cell(row=row_start + idx, column=7, value=sessao[4])  # Justificativa
        ws.cell(row=row_start + idx, column=8, value=sessao[5])  # Comentários

    os.makedirs("relatorios", exist_ok=True)
    wb.save(path_saida)
    return path_saida
